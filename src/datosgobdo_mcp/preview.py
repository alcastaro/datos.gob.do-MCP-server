"""Preview of resources: first/last/random N rows of CSV / XLSX / JSON.

datos.gob.do has no DataStore extension so we parse client-side. Stream with
a byte cap to avoid downloading huge files.
"""

from __future__ import annotations

import csv
import io
import json
import random
from typing import Any, Literal

import httpx

from .download import (
    PREVIEW_MAX_BYTES,
    _detect_encoding,
    classify_format,
    download_capped,
    err_text,
    looks_like_html,
)
from .netguard import NetGuardError

DEFAULT_ROWS = 20
MAX_ROWS = 200

SampleMode = Literal["head", "tail", "random"]


def _decode_text(data: bytes) -> tuple[str, str]:
    """Decode bytes to text. Returns (text, encoding_used)."""
    enc = _detect_encoding(data)
    try:
        return data.decode(enc), enc
    except UnicodeDecodeError:
        return data.decode("utf-8", errors="replace"), "utf-8 (with replacements)"


def _select_rows(
    all_rows: list[list[Any]],
    n: int,
    sample: SampleMode,
) -> list[list[Any]]:
    if not all_rows:
        return []
    if sample == "head":
        return all_rows[:n]
    if sample == "tail":
        return all_rows[-n:]
    if sample == "random":
        if len(all_rows) <= n:
            return all_rows
        return random.sample(all_rows, n)
    return all_rows[:n]


def _preview_csv(data: bytes, rows: int, sample: SampleMode) -> dict[str, Any]:
    text, encoding = _decode_text(data)
    sample_text = text[:4096]
    try:
        dialect = csv.Sniffer().sniff(sample_text, delimiters=",;\t|")
    except csv.Error:
        dialect = csv.excel
    reader = csv.reader(io.StringIO(text), dialect=dialect)
    try:
        header = next(reader)
    except StopIteration:
        return {"format": "csv", "error": "Archivo CSV vacío"}
    all_rows = list(reader)
    out_rows = _select_rows(all_rows, rows, sample)
    return {
        "format": "csv",
        "delimiter": dialect.delimiter,
        "encoding": encoding,
        "columns": header,
        "total_rows_in_download": len(all_rows),
        "rows_returned": len(out_rows),
        "sample_mode": sample,
        "rows": out_rows,
    }


def _jsonable(v: Any) -> Any:
    import datetime

    if isinstance(v, (datetime.datetime, datetime.date, datetime.time)):
        return v.isoformat()
    return v


def _preview_xlsx(data: bytes, rows: int, sample: SampleMode) -> dict[str, Any]:
    try:
        import openpyxl
    except ImportError:
        return {"format": "xlsx", "error": "openpyxl no instalado"}
    try:
        wb = openpyxl.load_workbook(io.BytesIO(data), read_only=True, data_only=True)
    except Exception as e:
        return {"format": "xlsx", "error": f"No se pudo abrir XLSX: {e}"}
    sheet = wb.active
    if sheet is None:
        wb.close()
        return {"format": "xlsx", "error": "Workbook sin hojas"}
    iterator = sheet.iter_rows(values_only=True)
    try:
        header_row = next(iterator)
    except StopIteration:
        wb.close()
        return {"format": "xlsx", "error": "Hoja vacía"}
    header = [str(c) if c is not None else "" for c in header_row]
    all_rows = [[_jsonable(c) for c in row] for row in iterator]
    out_rows = _select_rows(all_rows, rows, sample)
    sheets = wb.sheetnames
    wb.close()
    return {
        "format": "xlsx",
        "active_sheet": sheet.title,
        "all_sheets": sheets,
        "columns": header,
        "total_rows_in_download": len(all_rows),
        "rows_returned": len(out_rows),
        "sample_mode": sample,
        "rows": out_rows,
    }


def _preview_json(data: bytes, rows: int, sample: SampleMode) -> dict[str, Any]:
    text, _enc = _decode_text(data)
    try:
        obj = json.loads(text)
    except json.JSONDecodeError as e:
        return {"format": "json", "error": f"JSON inválido: {e}"}
    if isinstance(obj, list):
        selected = _select_rows(obj, rows, sample)  # type: ignore[arg-type]
        return {
            "format": "json-array",
            "total_items": len(obj),
            "rows_returned": len(selected),
            "sample_mode": sample,
            "rows": selected,
        }
    if isinstance(obj, dict):
        for key in ("data", "results", "items", "records"):
            inner = obj.get(key)
            if isinstance(inner, list):
                selected = _select_rows(inner, rows, sample)  # type: ignore[arg-type]
                return {
                    "format": "json-object",
                    "data_key": key,
                    "total_items": len(inner),
                    "rows_returned": len(selected),
                    "sample_mode": sample,
                    "other_keys": [k for k in obj.keys() if k != key],
                    "rows": selected,
                }
        return {"format": "json-object", "keys": list(obj.keys()), "data": obj}
    return {"format": "json-scalar", "value": obj}


async def _preview_via_cache(url: str, kind: str, rows: int, sample: SampleMode) -> dict[str, Any]:
    """Read a preview out of the Parquet cache instead of off the network.

    Imported lazily: preview is the cheap first-contact tool and importing
    DuckDB at module load would make the whole server slower to start.
    """
    from .analytics import AnalyticsError, ensure_cached

    try:
        parquet, meta = await ensure_cached(url, kind)
    except (httpx.HTTPError, AnalyticsError, NetGuardError, OSError) as e:
        return {"error": f"No se pudo cargar el recurso: {err_text(e)}"}

    import duckdb

    con = duckdb.connect(":memory:")
    try:
        p = str(parquet).replace("'", "''")
        described = con.execute(f"DESCRIBE SELECT * FROM read_parquet('{p}')").fetchall()
        columns = [r[0] for r in described]
        total = con.execute(f"SELECT COUNT(*) FROM read_parquet('{p}')").fetchone()[0]  # type: ignore[index]
        order = {
            "head": "",
            "tail": f"OFFSET {max(0, total - rows)}",
            "random": f"USING SAMPLE {rows} ROWS",
        }[sample]
        if sample == "random":
            sql = f"SELECT * FROM read_parquet('{p}') {order}"
        else:
            sql = f"SELECT * FROM read_parquet('{p}') LIMIT {rows} {order}"
        out_rows = [list(r) for r in con.execute(sql).fetchall()]
    except duckdb.Error as e:
        return {"error": f"No se pudo leer el recurso: {e}"}
    finally:
        con.close()

    return {
        "format": kind,
        "columns": columns,
        "total_rows_in_download": total,
        "rows_returned": len(out_rows),
        "sample_mode": sample,
        "rows": [[_jsonable(v) for v in r] for r in out_rows],
        "source": "parquet-cache",
        "cache": meta.get("cache"),
    }


async def preview_resource_data(
    url: str,
    fmt: str | None,
    rows: int = DEFAULT_ROWS,
    sample: SampleMode = "head",
) -> dict[str, Any]:
    """Download a resource and return a preview slice.

    Args:
        url: Direct URL to the file (from resource.url in CKAN).
        fmt: Declared format in CKAN (csv, xlsx, json, ods, pdf...). ODS is not
            previewable here — the analytics tools (get_resource_schema, etc.)
            do support it.
        rows: Rows to return (cap MAX_ROWS).
        sample: 'head' (first N), 'tail' (last N), or 'random' (uniform sample
            from the downloaded portion — biased if file was truncated by cap).

    Returns:
        Dict with parsed preview or {"error": ...}.
    """
    rows = min(max(int(rows), 1), MAX_ROWS)
    kind = classify_format(fmt)
    if kind is None:
        return {
            "error": f"Formato '{fmt}' no soportado para preview",
            "supported": ["CSV", "TSV", "XLSX", "ODS", "JSON"],
            "hint": "Descargar manualmente desde la URL del recurso.",
        }

    if kind == "ods":
        # ODS is a zip of XML, so there are no raw rows to slice: it has to be
        # transcoded before anything can read it. Rather than refuse — which is
        # what this tool used to do, and ODS is about a third of this catalog —
        # go through the cached analytics path, which already knows how.
        return await _preview_via_cache(url, kind, rows, sample)

    # If the analytics tools already pulled this resource, reuse their copy.
    # The audit measured the alternative: preview re-downloaded on every call,
    # which made it 20-25x slower than every other tool (p50 0.77s against
    # 0.03s) and put a fresh request on a government portal each time an
    # assistant glanced at a file it had already read.
    from .analytics import warm_parquet_for

    if warm_parquet_for(url) is not None:
        return await _preview_via_cache(url, kind, rows, sample)

    try:
        data, truncated = await download_capped(url, PREVIEW_MAX_BYTES)
    except httpx.HTTPStatusError as e:
        return {"error": f"HTTP {e.response.status_code} al bajar el recurso"}
    except httpx.HTTPError as e:
        return {"error": f"Error de red bajando recurso: {err_text(e)}"}
    except NetGuardError as e:
        return {"error": err_text(e)}

    if looks_like_html(data):
        return {
            "error": (
                "La URL devolvió una página HTML, no un archivo de datos — el portal "
                "probablemente respondió con una página web (HTTP 200) a un enlace de "
                "descarga caído o restringido."
            ),
            "hint": "Abrir la URL del recurso en un navegador para confirmar.",
        }

    if kind in ("csv", "tsv"):
        out = _preview_csv(data, rows, sample)
    elif kind in ("xlsx", "xls", "xlsm"):
        out = _preview_xlsx(data, rows, sample)
    else:  # json
        out = _preview_json(data, rows, sample)

    out["source_url"] = url
    out["bytes_downloaded"] = len(data)
    out["download_truncated"] = truncated
    if truncated and sample == "tail":
        out["warning"] = (
            "File exceeded preview byte cap (5 MB). 'tail' returns the tail of "
            "the downloaded portion, NOT the true file tail. Use schema/summarize "
            "tools which support a higher cap."
        )
    return out
